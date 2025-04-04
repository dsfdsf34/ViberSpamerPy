import time
import threading
import json
import os
import subprocess
import re
import keyboard
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from urllib.parse import unquote
from queue import Queue

# Путь для файла Excel с активными ссылками
excel_file = "active_links.xlsx"
# Путь для файла текстового документа с неактивными ссылками
inactive_links_file = "inactive_links.txt"
# Путь для сохранения настроек
settings_file = "settings.json"
pause_flag = False

def kill_chromedriver_process():
    try:
        # Пытаемся найти процессы chromedriver
        result = subprocess.run(['tasklist'], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)

        # Проверяем наличие chromedriver в списке запущенных процессов
        if 'chromedriver.exe' in result.stdout:
            subprocess.run(['taskkill', '/f', '/im', 'chromedriver.exe'], check=True)
            print("Процесс chromedriver успешно завершен.")
        else:
            print("Процесс chromedriver не найден.")
    except subprocess.CalledProcessError as e:
        print(f"Ошибка при попытке завершить процесс: {e}")
    except Exception as e:
        print(f"Неизвестная ошибка: {e}")

# Глобальные объекты Lock для синхронизации записи в файлы
excel_lock = threading.Lock()
text_lock = threading.Lock()

# Очередь для безопасной передачи данных между потоками
data_queue = Queue()


def toggle_pause():
    """Переключение паузы с помощью клавиши F8"""
    global pause_flag
    while True:
        if keyboard.is_pressed('F8'):
            pause_flag = not pause_flag
            print("Пауза!" if pause_flag else "Продолжение...")
            time.sleep(1)


def read_settings():
    """Чтение настроек из JSON-файла"""
    if os.path.exists(settings_file):
        with open(settings_file, "r", encoding="utf-8") as file:
            try:
                settings = json.load(file)
                return settings.get("path_to_excel"), settings.get("sheet_name"), settings.get("start_row"), settings.get("end_row")
            except json.JSONDecodeError:
                print("Ошибка в формате JSON.")
                return None, None, None, None
    else:
        print("Файл настроек не найден.")
        return None, None, None, None


def save_settings(path_to_excel, sheet_name, start_row, end_row):
    """Сохранение настроек в файл JSON"""
    settings = {"path_to_excel": path_to_excel, "sheet_name": sheet_name, "start_row": start_row, "end_row": end_row}
    with open(settings_file, "w", encoding="utf-8") as file:
        json.dump(settings, file, ensure_ascii=False, indent=4)
    print("Настройки сохранены.")

def get_sheet_names(path_to_excel):
    """Получение списка листов из Excel-файла"""
    wb = load_workbook(path_to_excel, read_only=True)
    return wb.sheetnames

def print_status(message):
    """Функция для вывода текущего состояния"""
    print(f"[{time.strftime('%H:%M:%S')}] {message}")

def create_excel_file():
    """Создание нового Excel файла для активных ссылок"""
    wb = Workbook()
    active_sheet = wb.active
    active_sheet.title = "Активные ссылки"
    active_sheet.append(["Ссылка", "Название группы", "Количество участников"])
    wb.save(excel_file)


def save_to_excel(sheet_name, link, group_name, members_count):
    """Сохранение активных ссылок и названий групп в файл Excel"""
    data_queue.put(('excel', sheet_name, link, group_name, members_count))  # Помещаем данные в очередь

def save_to_text(link):
    """Сохранение неактивных ссылок в текстовый файл"""
    with text_lock:  # Захватываем блокировку для записи в текстовый файл
        with open(inactive_links_file, "a", encoding="utf-8") as file:
            file.write(link + "\n")


def process_queue():
    """Обработка данных из очереди для записи в файлы с буферизацией"""
    buffer = []
    buffer_size = 100  # Размер буфера перед записью в файл
    print_status("Запуск обработки очереди...")

    while True:
        item = data_queue.get()
        if item[0] == 'excel':
            # Сохраняем данные в буфер
            buffer.append(item[1:])
            if len(buffer) >= buffer_size:
                flush_buffer_to_excel(buffer)
                buffer.clear()
        data_queue.task_done()

        # Сигнал завершения работы
        if item == "STOP":
            # Сохраняем оставшиеся данные из буфера
            if buffer:
                flush_buffer_to_excel(buffer)
            print_status("Обработка записи в таблицу завершена.")
            break


def flush_buffer_to_excel(buffer):
    """Записывает данные из буфера в Excel"""
    with excel_lock:  # Захватываем блокировку для записи в Excel
        wb = load_workbook(excel_file)
        sheet = wb.active
        for sheet_name, link, group_name, members_count in buffer:
            sheet.append([link, group_name, members_count])
        wb.save(excel_file)


def get_members_count(driver, url):
    """Извлекает количество участников с веб-страницы"""
    driver.get(url)
    time.sleep(1)  # Ждем загрузки страницы

    # Попытка получить количество участников из первого варианта
    try:
        # Находим содержимое тега <script> с типом __PREACT_CLI_DATA__
        script_content = driver.find_element(By.XPATH, '//script[@type="__PREACT_CLI_DATA__"]').get_attribute("textContent")

        # Декодируем URL-encoded и парсим JSON
        decoded_content = unquote(script_content)
        json_data = json.loads(decoded_content)

        # Извлекаем количество участников
        members_count = json_data.get("preRenderData", {}).get("members")
        if members_count:
            return members_count
    except Exception as e:
        print(f"Ошибка извлечения участников 1: {str(e).split(':', 1)[0]}")

    # Попытка получить количество участников из второго варианта
    try:
        members_count_xpath = driver.find_element(By.XPATH,
                                                  '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section[2]/div[1]/div[1]/app-account-info/div/ul/li[1]').text.strip()

        # Извлекаем только цифры из строки
        members_count_str = re.sub(r'\D', '', members_count_xpath)  # Удаляем все, кроме цифр

        # Преобразуем строку в число (целое число)
        if members_count_str:
            return int(members_count_str)  # Преобразуем строку в целое число
    except Exception as e:
        print(f"Ошибка извлечения участников 2: {str(e).split(':', 1)[0]}")

    # Попытка получить количество участников из третьего варианта
    try:
        # Новый вариант XPath
        members_count_xpath_2 = driver.find_element(By.XPATH,
                                                    '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section[2]/div[1]/div[1]/app-account-info/div/ul/li[2]').text.strip()

        # Извлекаем только цифры из строки
        members_count_str_2 = re.sub(r'\D', '', members_count_xpath_2)  # Удаляем все, кроме цифр

        # Преобразуем строку в число (целое число)
        if members_count_str_2:
            return int(members_count_str_2)  # Преобразуем строку в целое число
    except Exception as e:
        print(f"Ошибка извлечения участников 3: {str(e).split(':', 1)[0]}")

    # Попытка получить количество участников из 4 варианта
    try:
        # Новый вариант XPath
        members_count_xpath_3 = driver.find_element(By.XPATH,
                                                    '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section/div[1]/div[1]/app-account-info/div/ul/li[1]').text.strip()

        # Извлекаем только цифры из строки
        members_count_str_3 = re.sub(r'\D', '', members_count_xpath_3)  # Удаляем все, кроме цифр

        # Преобразуем строку в число (целое число)
        if members_count_str_3:
            return int(members_count_str_3)  # Преобразуем строку в целое число
    except Exception as e:
        print(f"Ошибка извлечения участников 4: {str(e).split(':', 1)[0]}")

    # Попытка получить количество участников из 5 варианта
    try:
        # Новый вариант XPath
        members_count_xpath_4 = driver.find_element(By.XPATH,
                                                    '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section/div[1]/div[1]/app-account-info/div/ul/li[2]').text.strip()

        # Извлекаем только цифры из строки
        members_count_str_3 = re.sub(r'\D', '', members_count_xpath_4)  # Удаляем все, кроме цифр

        # Преобразуем строку в число (целое число)
        if members_count_str_3:
            return int(members_count_str_3)  # Преобразуем строку в целое число
    except Exception as e:
        print(f"Ошибка извлечения участников 5: {str(e).split(':', 1)[0]}")

    # Если все варианты не сработали
    print("Все попытки извлечения участников не удались. 'Неизвестно'.")
    return "Неизвестно"



def check_viber_group_status(link, driver, invalid_texts):
    """Проверка статуса группы по ссылке и извлечение названия группы и количества участников"""
    try:
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[-1])
        time.sleep(1)
        driver.get(link)

        try:
            group_name_1 = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[1]/div[2]/h2').text
        except NoSuchElementException:
            group_name_1 = None

        try:
            group_name_2 = driver.find_element(By.XPATH,
                                               '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section[2]/div[1]/div[1]/div[1]/h2').text
        except NoSuchElementException:
            group_name_2 = None

        try:
            group_name_3 = driver.find_element(By.XPATH,
                                               '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section/div[1]/div[1]/div[1]/h2').text
        except NoSuchElementException:
            group_name_3 = None

        if group_name_1:
            group_name = group_name_1
        elif group_name_2:
            group_name = group_name_2
        elif group_name_3:
            group_name = group_name_3
        else:
            group_name = "Неизвестное название"

        # Извлекаем количество участников
        members_count = get_members_count(driver, link)
        time.sleep(1)

        page_source = driver.page_source
        if any(text in page_source for text in invalid_texts):
            return "Неактивна", group_name, members_count
        else:
            return "Активна", group_name, members_count
    except NoSuchElementException as e:
        print(f"Ошибка: элемент не найден - {e}")
        return "Ошибка", "Неизвестное название", "Неизвестно"
    except TimeoutException as e:
        print(f"Ошибка: время ожидания истекло - {e}")
        return "Ошибка", "Неизвестное название", "Неизвестно"
    except WebDriverException as e:
        print(f"Ошибка с WebDriver: {e}")
        return "Ошибка", "Неизвестное название", "Неизвестно"
    except Exception as e:
        print(f"Неизвестная ошибка: {e}")
        return f"Ошибка: {e}", "Неизвестное название", "Неизвестно"
    finally:
        driver.close()
        driver.switch_to.window(driver.window_handles[0])


def process_links(start_row, end_row, path_to_excel, sheet_name, invalid_texts):
    """Обработка ссылок в указанном диапазоне строк"""
    print_status(f"Начало обработки ссылок с {start_row} по {end_row} в листе '{sheet_name}'...")
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # Запуск в фоновом режиме
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
    time.sleep(1)  # Задержка на 2 секунды после запуска драйвера

    wb = load_workbook(path_to_excel, read_only=True)
    sheet = wb[sheet_name]

    for row in range(start_row, end_row + 1):
        link = sheet.cell(row=row, column=1).value
        status, group_name, members_count = check_viber_group_status(link, driver, invalid_texts)

        if status == "Активна":
            save_to_excel("Активные ссылки", link, group_name, members_count)
        else:
            save_to_text(link)

        print_status(f"Обработана ссылка: {link}, Статус: {status}, Группа: {group_name}, Участников: {members_count}")

    driver.quit()


def get_total_rows(path_to_excel, sheet_name):
    """Получение общего количества строк в Excel"""
    wb = load_workbook(path_to_excel, read_only=True)
    sheet = wb[sheet_name]
    return sheet.max_row


def get_rows_per_thread(start_row, end_row, num_threads):
    """Разделение строк между потоками"""
    total_rows = end_row - start_row + 1
    rows_per_thread = total_rows // num_threads
    remaining_rows = total_rows % num_threads

    thread_ranges = []
    current_row = start_row

    for i in range(num_threads):
        end_row_for_thread = current_row + rows_per_thread - 1
        if remaining_rows > 0:
            end_row_for_thread += 1
            remaining_rows -= 1
        thread_ranges.append((current_row, end_row_for_thread))
        current_row = end_row_for_thread + 1

    return thread_ranges


def main():
    """Основная функция скрипта"""
    path_to_excel, sheet_name, start_row, end_row = read_settings()
    use_saved_settings = input("Использовать сохраненные настройки? (y/n): ").lower()

    if use_saved_settings != "y":
        path_to_excel = input("Введите путь к файлу Excel с ссылками: ")
        sheet_names = get_sheet_names(path_to_excel)
        print(f"Доступные листы: {', '.join(sheet_names)}")
        sheet_name = input(f"Введите имя листа (по умолчанию '{sheet_names[0]}'): ") or sheet_names[0]
        start_row = int(input("С какой строки начать проверку? (по умолчанию 0): ") or 0)
        total_rows = get_total_rows(path_to_excel, sheet_name)
        end_row = int(input(f"До какой строки проверять? (по умолчанию {total_rows}): ") or total_rows)
        save_settings(path_to_excel, sheet_name, start_row, end_row)
    else:
        path_to_excel = input(
            f"Введите путь к файлу Excel с ссылками (текущая настройка: {path_to_excel}): ") or path_to_excel
        sheet_names = get_sheet_names(path_to_excel)
        print(f"Доступные листы: {', '.join(sheet_names)}")
        sheet_name = input(f"Введите имя листа (текущая настройка: {sheet_name}): ") or sheet_name
        start_row = int(input(f"С какой строки начать проверку? (текущая настройка: {start_row}): ") or start_row)
        total_rows = get_total_rows(path_to_excel, sheet_name)
        end_row = int(
            input(f"До какой строки проверять? (текущая настройка: {end_row}, по умолчанию {total_rows}): ") or end_row)

    num_threads = int(input("Введите количество потоков (по умолчанию 3): ") or 3)

    invalid_texts = [
        "Ссылка приглашения неактивна",
        "Срок действия приглашения истек",
        "Группа не существует",
        "Страница не найдена",
        "Ссылка не активна",
        "Ссылка неактивна",
        "Ссылка не найдена",
        "Группа не найдена"
    ]

    if not os.path.exists(excel_file):
        create_excel_file()

    # Запуск потока для обработки данных из очереди
    queue_thread = threading.Thread(target=process_queue, daemon=True)
    queue_thread.start()

    thread_ranges = get_rows_per_thread(start_row, end_row, num_threads)
    threads = []

    for i, (thread_start_row, thread_end_row) in enumerate(thread_ranges):
        thread = threading.Thread(target=process_links,
                                  args=(thread_start_row, thread_end_row, path_to_excel, sheet_name, invalid_texts))
        thread.start()
        threads.append(thread)

        # Добавляем паузу между инициализацией потоков
        time.sleep(1)

    for thread in threads:
        thread.join()

    # Посылаем сигнал завершения для потока записи
    data_queue.put("STOP")

    # Ожидаем завершения обработки очереди
    queue_thread.join()

    # Добавление вывода о завершении обработки с использованием print_status
    print_status(f"Парсинг завершен для листа: {sheet_name}.")


if __name__ == "__main__":
    # Вызов функции для завершения процесса chromedriver в самом начале
    kill_chromedriver_process()

    # Запуск потока для паузы
    pause_thread = threading.Thread(target=toggle_pause, daemon=True)
    pause_thread.start()

    main()
