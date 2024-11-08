import time
import keyboard
import os
import json
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

print("Нажмите F8 для закрытия текущей вкладки и открытия следующей")

# Путь к файлу настроек
settings_file = "settings.json"

def read_settings():
    """Чтение настроек из файла JSON."""
    if os.path.exists(settings_file):
        with open(settings_file, "r", encoding="utf-8") as file:
            try:
                settings = json.load(file)
                return settings.get("path_to_excel"), settings.get("start_row")
            except json.JSONDecodeError:
                print("Ошибка в формате JSON в файле настроек.")
                return None, None
    else:
        print("Файл настроек не найден.")
        return None, None

def save_settings(path_to_excel, start_row):
    """Сохранение настроек в файл JSON."""
    settings = {
        "path_to_excel": path_to_excel,
        "start_row": start_row
    }
    with open(settings_file, "w", encoding="utf-8") as file:
        json.dump(settings, file, ensure_ascii=False, indent=4)
    print("Настройки сохранены.")

def main():
    # Запрос на использование файла настроек
    use_settings = input("Хотите использовать файл настроек? (y/n): ").lower()

    if use_settings == "y":
        path_to_excel, start_row = read_settings()
        if path_to_excel is None or start_row is None:
            print("Настройки не найдены. Укажите путь и начальную строку вручную.")
            path_to_excel = input("Введите путь к файлу Excel с ссылками: ")
            start_row = int(input("С какой строки начать проверку? (по умолчанию 1): ") or 1)
            save_settings(path_to_excel, start_row)
        else:
            print(f"Используем настройки: путь к файлу Excel - {path_to_excel}, начало с строки {start_row}")
            start_row_input = input(f"Использовать строку {start_row} как начальную? Если нет, введите другую строку: ")
            start_row = int(start_row_input) if start_row_input else start_row
            save_settings(path_to_excel, start_row)
    else:
        path_to_excel = input("Введите путь к файлу Excel с ссылками: ")
        start_row = int(input("С какой строки начать проверку? (по умолчанию 1): ") or 1)
        save_settings(path_to_excel, start_row)

    # Загрузка файла Excel
    try:
        workbook = load_workbook(path_to_excel)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка при загрузке файла Excel: {e}")
        return

    # Инициализация драйвера
    try:
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))
    except Exception as e:
        print(f"Ошибка при запуске браузера: {e}")
        return

    # Открытие ссылок из Excel
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_col=1, values_only=True), start=start_row):
        link = row[0]
        if link:
            print(f"Открытие ссылки из строки {row_idx}: {link}")
            driver.execute_script("window.open('');")  # Открываем новую вкладку
            driver.switch_to.window(driver.window_handles[-1])  # Переход на новую вкладку
            driver.get(link)  # Открываем ссылку
            time.sleep(2)  # Небольшая пауза для загрузки страницы

            # Ждем нажатия клавиши F8, чтобы закрыть текущую вкладку и открыть следующую ссылку
            print("Нажмите F8 для перехода к следующей ссылке...")
            while not keyboard.is_pressed('F8'):
                time.sleep(0.1)  # Пауза для уменьшения нагрузки на CPU

            # Закрываем текущую вкладку и переключаемся на предыдущую (если она есть)
            driver.close()
            driver.switch_to.window(driver.window_handles[0])

    driver.quit()  # Закрытие браузера после завершения работы
    print("Открытие ссылок завершено.")

if __name__ == "__main__":
    main()
