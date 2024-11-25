import time
import threading
import json
import os
import re
import keyboard
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException
from urllib.parse import unquote
from queue import Queue

# Путь для файла Excel с активными ссылками
excel_file = "active_links.xlsx"
# Путь для файла текстового документа с неактивными ссылками
inactive_links_file = "inactive_links.txt"
# Путь для сохранения настроек
settings_file = "settings.json"
pause_flag = False

# Глобальные объекты Lock для синхронизации записи в файлы
excel_lock = threading.Lock()
text_lock = threading.Lock()

# Очередь для безопасной передачи данных между потоками
data_queue = Queue()


def toggle_pause():
    """Переключение паузы с помощью клавиши F8"""
    global pause_flag
    while True:
        if keyboard.is_pressed('F8'):
            pause_flag = not pause_flag
            print("Пауза!" if pause_flag else "Продолжение...")
            time.sleep(1)


def read_settings():
    """Чтение настроек из JSON-файла"""
    if os.path.exists(settings_file):
        with open(settings_file, "r", encoding="utf-8") as file:
            try:
                settings = json.load(file)
                return settings.get("path_to_excel"), settings.get("start_row"), settings.get("end_row")
            except json.JSONDecodeError:
                print("Ошибка в формате JSON.")
                return None, None, None
    else:
        print("Файл настроек не найден.")
        return None, None, None


def save_settings(path_to_excel, start_row, end_row):
    """Сохранение настроек в файл JSON"""
    settings = {"path_to_excel": path_to_excel, "start_row": start_row, "end_row": end_row}
    with open(settings_file, "w", encoding="utf-8") as file:
        json.dump(settings, file, ensure_ascii=False, indent=4)
    print("Настройки сохранены.")


def create_excel_file():
    """Создание нового Excel файла для активных ссылок"""
    wb = Workbook()
    active_sheet = wb.active
    active_sheet.title = "Активные ссылки"
    active_sheet.append(["Ссылка", "Название группы", "Количество участников"])
    wb.save(excel_file)


def save_to_excel(sheet_name, link, group_name, members_count):
    """Сохранение активных ссылок и названий групп в файл Excel"""
    data_queue.put(('excel', sheet_name, link, group_name, members_count))  # Помещаем данные в очередь


def save_to_text(link):
    """Сохранение неактивных ссылок в текстовый файл"""
    with text_lock:  # Захватываем блокировку для записи в текстовый файл
        with open(inactive_links_file, "a", encoding="utf-8") as file:
            file.write(link + "\n")


def process_queue():
    """Обработка данных из очереди для записи в файлы"""
    while True:
        item = data_queue.get()
        if item[0] == 'excel':
            sheet_name, link, group_name, members_count = item[1], item[2], item[3], item[4]
            with excel_lock:  # Захватываем блокировку для записи в Excel
                wb = load_workbook(excel_file)
                sheet = wb[sheet_name]
                sheet.append([link, group_name, members_count])
                wb.save(excel_file)
        data_queue.task_done()


def get_members_count(driver, url):
    """Извлекает количество участников с веб-страницы"""
    driver.get(url)
    time.sleep(5)  # Ждем загрузки страницы

    # Попытка получить количество участников из первого варианта
    try:
        # Находим содержимое тега <script> с типом __PREACT_CLI_DATA__
        script_content = driver.find_element(By.XPATH, '//script[@type="__PREACT_CLI_DATA__"]').get_attribute("textContent")

        # Декодируем URL-encoded и парсим JSON
        decoded_content = unquote(script_content)
        json_data = json.loads(decoded_content)

        # Извлекаем количество участников
        members_count = json_data.get("preRenderData", {}).get("members")
        if members_count:
            return members_count
    except Exception as e:
        print(f"Ошибка при извлечении количества участников из первого варианта: {e}")

    # Попытка получить количество участников из второго варианта
    try:
        members_count_xpath = driver.find_element(By.XPATH,
                                                  '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section[2]/div[1]/div[1]/app-account-info/div/ul/li[1]').text.strip()

        # Извлекаем только цифры из строки
        members_count_str = re.sub(r'\D', '', members_count_xpath)  # Удаляем все, кроме цифр

        # Преобразуем строку в число (целое число)
        if members_count_str:
            return int(members_count_str)  # Преобразуем строку в целое число
    except Exception as e:
        print(f"Ошибка при извлечении количества участников из второго варианта: {e}")

    # Попытка получить количество участников из третьего варианта
    try:
        # Новый вариант XPath
        members_count_xpath_2 = driver.find_element(By.XPATH,
                                                    '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section[2]/div[1]/div[1]/app-account-info/div/ul/li[2]').text.strip()

        # Извлекаем только цифры из строки
        members_count_str_2 = re.sub(r'\D', '', members_count_xpath_2)  # Удаляем все, кроме цифр

        # Преобразуем строку в число (целое число)
        if members_count_str_2:
            return int(members_count_str_2)  # Преобразуем строку в целое число
    except Exception as e:
        print(f"Ошибка при извлечении количества участников из третьего варианта: {e}")
        return "Неизвестно"

    # Если все три варианта не сработали
    return "Неизвестно"



def check_viber_group_status(link, driver, invalid_texts):
    """Проверка статуса группы по ссылке и извлечение названия группы и количества участников"""
    try:
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[-1])
        driver.get(link)
        time.sleep(2)

        try:
            group_name_1 = driver.find_element(By.XPATH, '/html/body/div/div[2]/div[1]/div[2]/h2').text
        except NoSuchElementException:
            group_name_1 = None

        try:
            group_name_2 = driver.find_element(By.XPATH,
                                               '/html/body/app-root/vbr-page/vbr-content/app-main/app-account/article/section[2]/div[1]/div[1]/div[1]/h2').text
        except NoSuchElementException:
            group_name_2 = None

        group_name = group_name_1 if group_name_1 else (group_name_2 if group_name_2 else "Неизвестное название")

        # Извлекаем количество участников
        members_count = get_members_count(driver, link)

        page_source = driver.page_source
        if any(text in page_source for text in invalid_texts):
            return "Неактивна", group_name, members_count
        else:
            return "Активна", group_name, members_count
    except NoSuchElementException as e:
        print(f"Ошибка: элемент не найден - {e}")
        return "Ошибка", "Неизвестное название", "Неизвестно"
    except TimeoutException as e:
        print(f"Ошибка: время ожидания истекло - {e}")
        return "Ошибка", "Неизвестное название", "Неизвестно"
    except WebDriverException as e:
        print(f"Ошибка с WebDriver: {e}")
        return "Ошибка", "Неизвестное название", "Неизвестно"
    except Exception as e:
        print(f"Неизвестная ошибка: {e}")
        return f"Ошибка: {e}", "Неизвестное название", "Неизвестно"
    finally:
        driver.close()
        driver.switch_to.window(driver.window_handles[0])


def process_links(start_row, end_row, path_to_excel, invalid_texts):
    """Обработка ссылок в указанном диапазоне строк"""
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")  # Запуск в фоновом режиме
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

    wb = load_workbook(path_to_excel)
    sheet = wb.active

    for row in range(start_row, end_row + 1):
        link = sheet.cell(row=row, column=1).value
        status, group_name, members_count = check_viber_group_status(link, driver, invalid_texts)

        if status == "Активна":
            save_to_excel("Активные ссылки", link, group_name, members_count)
        else:
            save_to_text(link)

    driver.quit()


def get_total_rows(path_to_excel):
    """Получение общего количества строк в Excel"""
    wb = load_workbook(path_to_excel)
    sheet = wb.active
    return sheet.max_row


def get_rows_per_thread(start_row, end_row, num_threads):
    """Разделение строк между потоками"""
    total_rows = end_row - start_row + 1
    rows_per_thread = total_rows // num_threads
    remaining_rows = total_rows % num_threads

    thread_ranges = []
    current_row = start_row

    for i in range(num_threads):
        end_row_for_thread = current_row + rows_per_thread - 1
        if remaining_rows > 0:
            end_row_for_thread += 1
            remaining_rows -= 1
        thread_ranges.append((current_row, end_row_for_thread))
        current_row = end_row_for_thread + 1

    return thread_ranges


def main():
    """Основная функция скрипта"""
    path_to_excel, start_row, end_row = read_settings()
    use_saved_settings = input("Использовать сохраненные настройки? (y/n): ").lower()

    if use_saved_settings != "y":
        path_to_excel = input("Введите путь к файлу Excel с ссылками: ")
        start_row = int(input("С какой строки начать проверку? (по умолчанию 0): ") or 0)
        total_rows = get_total_rows(path_to_excel)
        end_row = int(input(f"До какой строки проверять? (по умолчанию {total_rows}): ") or total_rows)
        save_settings(path_to_excel, start_row, end_row)
    else:
        path_to_excel = input(
            f"Введите путь к файлу Excel с ссылками (текущая настройка: {path_to_excel}): ") or path_to_excel
        start_row = int(input(f"С какой строки начать проверку? (текущая настройка: {start_row}): ") or start_row)
        total_rows = get_total_rows(path_to_excel)
        end_row = int(
            input(f"До какой строки проверять? (текущая настройка: {end_row}, по умолчанию {total_rows}): ") or end_row)

    num_threads = int(input("Введите количество потоков (по умолчанию 3): ") or 3)

    invalid_texts = [
        "Ссылка приглашения неактивна",
        "Срок действия приглашения истек",
        "Группа не существует",
        "Страница не найдена",
        "Ссылка не активна",
        "Ссылка неактивна",
        "Ссылка не найдена",
        "Группа не найдена"
    ]

    if not os.path.exists(excel_file):
        create_excel_file()

    # Запуск потока для обработки данных из очереди
    queue_thread = threading.Thread(target=process_queue, daemon=True)
    queue_thread.start()

    thread_ranges = get_rows_per_thread(start_row, end_row, num_threads)
    threads = []

    for i, (thread_start_row, thread_end_row) in enumerate(thread_ranges):
        thread = threading.Thread(target=process_links,
                                  args=(thread_start_row, thread_end_row, path_to_excel, invalid_texts))
        thread.start()
        threads.append(thread)

    for thread in threads:
        thread.join()

    # Ожидаем завершения обработки очереди
    data_queue.join()

    print("Обработка завершена.")


if __name__ == "__main__":
    # Запуск потока для паузы
    pause_thread = threading.Thread(target=toggle_pause, daemon=True)
    pause_thread.start()

    main()
