import time
import keyboard
import threading
import json
import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

# Путь для неактивных ссылок
inactive_links_file = "inactive_links.txt"

# Путь для активных ссылок
active_links_file = "active_links.txt"

# Настройки
settings_file = "settings.json"
pause_flag = False


def toggle_pause():
    """Переключение паузы с помощью клавиши F8"""
    global pause_flag
    while True:
        if keyboard.is_pressed('F8'):
            pause_flag = not pause_flag
            print("Пауза!" if pause_flag else "Продолжение...")
            time.sleep(1)


def read_settings():
    """Чтение настроек из JSON-файла"""
    if os.path.exists(settings_file):
        with open(settings_file, "r", encoding="utf-8") as file:
            try:
                settings = json.load(file)
                return settings.get("path_to_excel"), settings.get("start_row"), settings.get("end_row")
            except json.JSONDecodeError:
                print("Ошибка в формате JSON.")
                return None, None, None
    else:
        print("Файл настроек не найден.")
        return None, None, None


def save_settings(path_to_excel, start_row, end_row):
    """Сохранение настроек в файл JSON"""
    settings = {"path_to_excel": path_to_excel, "start_row": start_row, "end_row": end_row}
    with open(settings_file, "w", encoding="utf-8") as file:
        json.dump(settings, file, ensure_ascii=False, indent=4)
    print("Настройки сохранены.")


def log_message(message, file_path):
    """Запись сообщений в файл активных или неактивных ссылок с номером строки"""
    with open(file_path, "a", encoding="utf-8") as f:
        f.write(message + "\n")


def check_viber_group_status(link, driver, invalid_texts):
    """Проверка статуса группы по ссылке"""
    try:
        driver.execute_script("window.open('');")
        driver.switch_to.window(driver.window_handles[-1])
        driver.get(link)
        time.sleep(2)

        page_source = driver.page_source
        if any(text in page_source for text in invalid_texts):
            return "Неактивна"
        else:
            return "Активна"
    except Exception as e:
        return f"Ошибка: {e}"
    finally:
        driver.close()
        driver.switch_to.window(driver.window_handles[0])


def process_links(start_row, end_row, path_to_excel, invalid_texts):
    """Обработка ссылок в Excel"""
    try:
        workbook = load_workbook(path_to_excel)
        sheet = workbook.active
    except Exception as e:
        print(f"Ошибка при загрузке файла Excel: {e}")
        return

    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

    # Процесс перебора строк с учетом start_row и end_row
    for row_idx, row in enumerate(sheet.iter_rows(min_row=start_row, max_row=end_row, max_col=1, values_only=True),
                                  start=start_row):
        link = row[0]
        if link:
            status = check_viber_group_status(link, driver, invalid_texts)
            print(f'Поток "{threading.current_thread().name}" Строка: "{row_idx}" {link} - Статус: {status}')

            # Записываем ссылку в соответствующий файл в зависимости от статуса
            if status == "Активна":
                log_message(f"{link}", active_links_file)
            else:
                log_message(f"{link}", inactive_links_file)

    driver.quit()


def get_total_rows(path_to_excel):
    """Получение общего количества строк в Excel"""
    try:
        workbook = load_workbook(path_to_excel)
        sheet = workbook.active
        return sheet.max_row
    except Exception as e:
        print(f"Ошибка при загрузке файла Excel: {e}")
        return 0


def get_rows_per_thread(start_row, end_row, num_threads):
    """Рассчитывает, сколько строк обрабатывать каждому потоку"""
    total_rows = end_row - start_row + 1
    rows_per_thread = total_rows // num_threads
    remaining_rows = total_rows % num_threads

    thread_ranges = []
    current_row = start_row
    for i in range(num_threads):
        end_row_for_thread = current_row + rows_per_thread + (1 if i < remaining_rows else 0) - 1
        thread_ranges.append((current_row, end_row_for_thread))
        current_row = end_row_for_thread + 1

    return thread_ranges


def main():
    """Основная функция скрипта"""
    # Переменные для пути и начальной строки
    path_to_excel = None
    start_row = None
    end_row = None  # Инициализируем end_row

    use_saved_settings = input("Использовать сохраненные настройки? (y/n): ").lower()
    if use_saved_settings == "y":
        path_to_excel, start_row, end_row = read_settings()
        if path_to_excel is None or start_row is None or end_row is None:
            print("Не удалось загрузить настройки из файла.")
            use_saved_settings = "n"

    if use_saved_settings == "n":
        path_to_excel = input("Введите путь к файлу Excel с ссылками: ")
        start_row = int(input("С какой строки начать проверку? (по умолчанию 0): ") or 0)

        total_rows = get_total_rows(path_to_excel)  # Получаем общее количество строк в таблице
        print(f"В таблице найдено {total_rows} строк.")  # Сообщаем количество строк

        # Устанавливаем максимальное количество строк по умолчанию
        default_end_row = total_rows
        end_row_input = input(f"До какой строки проверять? (по умолчанию {default_end_row}): ")

        # Если поле пустое, сообщаем, что будет использовано всё количество строк
        if not end_row_input:
            print(f"Будут использованы все строки (до {default_end_row}).")
            end_row = total_rows
        else:
            end_row = int(end_row_input)

        # Если указана строка, превышающая общее количество, то обработать все строки
        if end_row > total_rows:
            print(f"В таблице всего {total_rows} строк. Будем обрабатывать все строки.")
            end_row = total_rows

        save_settings(path_to_excel, start_row, end_row)
    else:
        total_rows = get_total_rows(path_to_excel)
        print(f"В таблице найдено {total_rows} строк.")  # Сообщаем количество строк

        # Если указана строка, превышающая общее количество, то обработать все строки
        if end_row > total_rows:
            print(f"В таблице всего {total_rows} строк. Будем обрабатывать все строки.")
            end_row = total_rows

        start_row = int(
            input(f"Введите начальную строку для продолжения (текущее значение {start_row}): ") or start_row)
        end_row = int(
            input(f"Введите конечную строку для продолжения (текущее значение {start_row + 100}): ") or end_row)

    num_threads = int(input("Введите количество потоков для проверки: "))

    # Запросить выбор стандартных или собственных слов
    choice = input("Использовать стандартные (для поиска неактивных групп) слова для поиска? (y/n): ").lower()
    if choice == "y":
        invalid_texts = [
            "Ссылка приглашения неактивна",
            "Срок действия приглашения истек",
            "Группа не существует",
            "Страница не найдена",
            "Ссылка не активна",
            "Ссылка неактивна",
            "Ссылка не найдена",
            "Группа не найдена"
        ]
    else:
        invalid_texts = input("Введите собственные слова для поиска (через запятую): ").split(',')

    # Получаем диапазоны строк для каждого потока
    thread_ranges = get_rows_per_thread(start_row, end_row, num_threads)

    pause_thread = threading.Thread(target=toggle_pause)
    pause_thread.daemon = True
    pause_thread.start()

    threads = []
    for i, (thread_start_row, thread_end_row) in enumerate(thread_ranges):
        thread = threading.Thread(target=process_links,
                                  args=(thread_start_row, thread_end_row, path_to_excel, invalid_texts))
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    print("Проверка завершена для всех потоков.")


if __name__ == "__main__":
    main()
